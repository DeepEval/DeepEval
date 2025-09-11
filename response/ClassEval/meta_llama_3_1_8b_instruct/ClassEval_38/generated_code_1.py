import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            data = list(sheet.rows)
            return data
        except FileNotFoundError:
            print(f"The file {file_name} does not exist.")
            return None
        except Exception as e:
            print(f"An error occurred: {e}")
            return None

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            for row in data:
                ws.append(row)
            wb.save(file_name)
            return 1
        except Exception as e:
            print(f"An error occurred: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        try:
            wb = openpyxl.load_workbook(save_file_name)
            sheet = wb.active
            for row in sheet.rows:
                if N <= len(row):
                    row[N - 1].value = str(row[N - 1].value).upper()
            output_file_name = save_file_name.split('.')[0] + '_processed.xlsx'
            wb.save(output_file_name)
            return self.write_excel([list(row) for row in sheet.rows], output_file_name), output_file_name
        except Exception as e:
            print(f"An error occurred: {e}")
            return None, None


if __name__ == "__main__":
    processor = ExcelProcessor()

    # Test case for read_excel method
    data = processor.read_excel('test_data.xlsx')
    if data is not None:
        print("Data read from Excel file:")
        for row in data:
            print(row)

    # Test case for write_excel method
    new_data = [
        ('Name', 'Age', 'Country'),
        ('John', 25, 'USA'),
        ('Alice', 30, 'Canada'),
        ('Bob', 35, 'Australia'),
        ('Julia', 28, 'Germany')
    ]
    result = processor.write_excel(new_data, 'test_data.xlsx')
    print(f"Write Excel result: {result}")

    # Test case for process_excel_data method
    success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
    if success:
        print(f"Processing successful, output file saved to: {output_file}")
    else:
        print("Processing failed.")