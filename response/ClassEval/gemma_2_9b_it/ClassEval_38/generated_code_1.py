import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows():
            row_data = [cell.value for cell in row]
            data.append(row_data)
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('Name', 'Age', 'Country'),
        >>>     ('John', 25, 'USA'),
        >>>     ('Alice', 30, 'Canada'),
        >>>     ('Bob', 35, 'Australia'),
        >>>     ('Julia', 28, 'Germany')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row_index, row_data in enumerate(data):
            for col_index, value in enumerate(row_data):
                sheet.cell(row=row_index + 1, column=col_index + 1).value = value
        workbook.save(file_name)
        return 1

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        workbook = openpyxl.load_workbook(save_file_name)
        sheet = workbook.active
        for row in sheet.iter_rows():
            row[N - 1].value = row[N - 1].value.upper()
        workbook.save(f"{save_file_name}_processed.xlsx")
        return 1, f"{save_file_name}_processed.xlsx"



if __name__ == "__main__":
    processor = ExcelProcessor()
    # Test case for read_excel
    data = processor.read_excel('test_data.xlsx')
    print(data)

    # Test case for write_excel
    new_data = [
        ('Name', 'Age', 'Country'),
        ('John', 25, 'USA'),
        ('Alice', 30, 'Canada'),
        ('Bob', 35, 'Australia'),
        ('Julia', 28, 'Germany')
    ]
    success = processor.write_excel(new_data, 'test_data.xlsx')
    print(success)

    # Test case for process_excel_data
    success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
    print(success)
    print(output_file)