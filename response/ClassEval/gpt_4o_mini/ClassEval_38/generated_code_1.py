import openpyxl

class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data,
    as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name: str, Excel file name to read
        :return: list of data, Data in Excel
        """
        data = []
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data
        except Exception as e:
            print(f"Error reading file: {e}")
            return []

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"Error writing file: {e}")
            return 0

    def process_excel_data(self, N, file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param file_name: str, source file name
        :return: (int, str), The former is the return value of write_excel,
                 while the latter is the saved file name of the processed data
        """
        data = self.read_excel(file_name)
        if not data:
            return 0, ""

        # Convert the specified column to uppercase
        for i in range(1, len(data)):  # Starting from 1 to skip the header row
            if len(data[i]) > N:
                data[i] = list(data[i])
                data[i][N] = str(data[i][N]).upper()

        output_file_name = "processed_" + file_name
        result = self.write_excel(data, output_file_name)
        return result, output_file_name


if __name__ == "__main__":
    processor = ExcelProcessor()
    
    # Test case for write_excel
    new_data = [
        ('Name', 'Age', 'Country'),
        ('John', 25, 'USA'),
        ('Alice', 30, 'Canada'),
        ('Bob', 35, 'Australia'),
        ('Julia', 28, 'Germany')
    ]
    write_result = processor.write_excel(new_data, 'test_data.xlsx')
    print(f"Write result: {write_result}")  # Expected output: 1

    # Test case for read_excel
    read_data = processor.read_excel('test_data.xlsx')
    print("Read data:")
    for row in read_data:
        print(row)  # Should print the contents of the Excel file

    # Test case for process_excel_data
    process_result, output_file = processor.process_excel_data(0, 'test_data.xlsx')
    print(f"Process result: {process_result}, Output file: {output_file}")  # Expected output: (1, 'processed_test_data.xlsx')