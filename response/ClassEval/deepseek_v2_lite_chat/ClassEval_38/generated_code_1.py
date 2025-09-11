import openpyxl

class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name: str, Excel file name to read
        :return: list of data, Data in Excel
        """
        try:
            workbook = openpyxl.load_workbook(file_name)
            data = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in range(1, sheet.max_row + 1):
                    row_data = []
                    for column in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row, column).value
                        row_data.append(cell_value)
                    data.append(row_data)
            return data
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return []

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            workbook = openpyxl.Workbook()
            for index, row in enumerate(data, start=1):
                workbook.create_sheet(title=str(index))
                sheet = workbook[str(index)]
                for col_num, cell_value in enumerate(row, start=1):
                    sheet.cell(row=1, column=col_num).value = cell_value
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"Error writing Excel file: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        try:
            data = self.read_excel(save_file_name)
            if data:
                for sheet_index, sheet in enumerate(data, start=1):
                    for row_index, row in enumerate(sheet, start=2):
                        row[N-1] = row[N-1].upper() if N <= len(row) else row[N-1]
                return self.write_excel(data, save_file_name)
            else:
                return 0, "No data to process"
        except Exception as e:
            print(f"Error processing Excel data: {e}")
            return 0, "Error processing data"

# Test cases
if __name__ == "__main__":
    processor = ExcelProcessor()
    data = [
        ('Name', 'Age', 'Country'),
        ('John', 25, 'USA'),
        ('Alice', 30, 'Canada'),
        ('Bob', 35, 'Australia'),
        ('Julia', 28, 'Germany')
    ]
    # Test read_excel
    print(processor.read_excel('test_data.xlsx'))

    # Test write_excel
    new_data = [
        ('John', 25, 'USA'),
        ('Alice', 30, 'Canada'),
        ('Bob', 35, 'Australia'),
        ('Julia', 28, 'Germany')
    ]
    print(processor.write_excel(new_data, 'new_test_data.xlsx'))

    # Test process_excel_data
    success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
    print(success, output_file)