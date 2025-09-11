import openpyxl

class ExcelProcessor:
    """
    This is a class for processing Excel files, including reading and writing Excel data,
    as well as processing specific operations and saving as a new Excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name: str, Excel file name to read
        :return: list of data, Data in Excel
        """
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"An error occurred: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change (0-based index)
        :param save_file_name: str, source file name
        :return: (int, str), The former is the return value of write_excel,
                while the latter is the saved file name of the processed data
        """
        data = self.read_excel(save_file_name)
        processed_data = []
        for row in data:
            row = list(row)
            if N < len(row) and isinstance(row[N], str):
                row[N] = row[N].upper()
            processed_data.append(tuple(row))

        output_file_name = "processed_" + save_file_name
        success = self.write_excel(processed_data, output_file_name)
        return success, output_file_name

if __name__ == "__main__":
    # Create an instance of ExcelProcessor
    processor = ExcelProcessor()
    
    # Test read_excel
    try:
        data = processor.read_excel('test_data.xlsx')
        print("Read Data:", data)
    except FileNotFoundError:
        print("File not found. Please ensure 'test_data.xlsx' exists for testing.")

    # Test write_excel
    new_data = [
        ('Name', 'Age', 'Country'),
        ('John', 25, 'USA'),
        ('Alice', 30, 'Canada'),
        ('Bob', 35, 'Australia'),
        ('Julia', 28, 'Germany')
    ]
    write_status = processor.write_excel(new_data, 'test_write.xlsx')
    print("Write Status:", "Success" if write_status == 1 else "Failed")
    
    # Test process_excel_data
    try:
        success, output_file = processor.process_excel_data(0, 'test_data.xlsx')
        print("Process Status:", "Success" if success == 1 else "Failed", "Output File:", output_file)
    except FileNotFoundError:
        print("File not found. Please ensure 'test_data.xlsx' exists for testing.")